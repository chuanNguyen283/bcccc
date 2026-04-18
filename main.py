import os
import shutil
from copy import copy

import telebot
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# File mẫu (logo + màu theme): đặt cùng thư mục với main.py hoặc chỉnh BC_TUAN_TEMPLATE
_BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(_BASE, "BC-TUAN02.04.26.xlsx")

# Token bot: ưu tiên biến môi trường TELEGRAM_BOT_TOKEN (an toàn hơn khi deploy)
TELEGRAM_BOT_TOKEN = os.environ.get(
    "TELEGRAM_BOT_TOKEN",
    "8482904890:AAFfobK0FOaDEIEUuySc_8SqPV890gOKy2o",
)
bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)

FONT_NAME = "Times New Roman"
FONT_SIZE = 11


def _to_naive_local(dt):
    """API trả ISO có thể có tz (UTC); datetime.now() là naive local — cần thống nhất."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt
    return dt.astimezone().replace(tzinfo=None)


def _parse_dt(date_str):
    if not date_str:
        return None
    try:
        dt = pd.to_datetime(date_str).to_pydatetime()
        return _to_naive_local(dt)
    except Exception:
        return None


def format_ts_part(date_str):
    """Giống mẫu: dd/mm/yyyy  hh:mm:ss (hai dấu cách giữa ngày và giờ)."""
    dt = _parse_dt(date_str)
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y  %H:%M:%S")


def format_ngay_cell(created_at, completed_at):
    a = format_ts_part(created_at)
    c = format_ts_part(completed_at)
    if a and c:
        return f"{a} {c} "
    if a:
        return f"Từ {a}  "
    return ""


def week_bounds(reference: datetime):
    """Tuần theo thứ Hai (0)–Chủ nhật, khớp logic lọc cũ."""
    monday = (reference - timedelta(days=reference.weekday())).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    sunday = monday + timedelta(days=6, hours=23, minutes=59, seconds=59)
    return monday, sunday


def in_week(created_at, week_start: datetime, week_end: datetime):
    dt = _parse_dt(created_at)
    if not dt:
        return False
    return week_start <= dt <= week_end


def priority_label(raw_priority):
    p = str(raw_priority or "").lower().strip()
    mapping = {
        "critical": "Tối Khẩn",
        "urgent": "Khẩn",
        "high": "Cao",
        "medium": "Trung Bình",
        "low": "Thấp",
    }
    return mapping.get(p, raw_priority or "")


def first_nonempty_str(*sources, keys):
    for src in sources:
        if not isinstance(src, dict):
            continue
        for k in keys:
            v = src.get(k)
            if isinstance(v, str) and v.strip():
                return v.strip()
    return ""


def _week_header_text(week_monday: datetime, week_sunday: datetime) -> str:
    iso_week = week_monday.isocalendar().week
    return (
        f"BÁO CÁO TUẦN {iso_week:02d} \n"
        f"(Từ {week_monday.strftime('%d/%m/%Y')} đến {week_sunday.strftime('%d/%m/%Y')})\n"
        f"BỘ PHẬN BÁO CÁO:  Phòng Vận Hành Dịch Vụ"
    )


def _apply_cell_style(src, dest):
    dest.font = copy(src.font)
    dest.border = copy(src.border)
    dest.fill = copy(src.fill)
    dest.alignment = copy(src.alignment)
    dest.number_format = src.number_format


def _embed_logo_from_template(ws, template_path: str) -> None:
    """Openpyxl làm mất ảnh khi save; chèn lại logo từ xl/media trong file mẫu (cần Pillow)."""
    try:
        import tempfile
        import zipfile

        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return
    try:
        with zipfile.ZipFile(template_path) as z:
            media = [
                n
                for n in z.namelist()
                if n.startswith("xl/media/")
                and n.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp"))
            ]
            if not media:
                return
            media.sort()
            td = tempfile.mkdtemp()
            name = media[0]
            z.extract(name, td)
            png = os.path.join(td, *name.split("/"))
            img = XLImage(png)
            ws.add_image(img, "C2")
    except Exception:
        pass


def _find_stats_title_row(ws, max_scan: int = 500):
    for r in range(1, min(ws.max_row + 1, max_scan)):
        v = ws.cell(r, 5).value
        if v and "BÁO CÁO THỐNG KÊ" in str(v).strip():
            return r
    return None


def _has_merge_E_to_G_one_row(ws, row: int) -> bool:
    for rng in ws.merged_cells.ranges:
        if (
            rng.min_row == rng.max_row == row
            and rng.min_col == 5
            and rng.max_col == 7
        ):
            return True
    return False


def _cleanup_orphan_EG_row_merges(ws, keep_stats_row: int) -> None:
    """Bỏ merge E:G dư (sau insert/delete openpyxl đôi khi để lại vùng trống, ví dụ E24:G24)."""
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row != rng.max_row:
            continue
        if rng.min_col != 5 or rng.max_col != 7:
            continue
        r = rng.min_row
        if r == keep_stats_row:
            continue
        v = ws.cell(r, 5).value
        if v is None or (isinstance(v, str) and not str(v).strip()):
            ws.unmerge_cells(str(rng))


def _ensure_bao_cao_thong_ke_merge(ws, stats_row: int) -> None:
    """Gộp 3 cột E:G cho dòng 'BÁO CÁO THỐNG KÊ' (sau insert/delete merge dễ bị vỡ)."""
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row != stats_row or rng.max_row != stats_row:
            continue
        if rng.min_col > 7 or rng.max_col < 5:
            continue
        if rng.min_col == 5 and rng.max_col == 7:
            continue
        ws.unmerge_cells(str(rng))
    if not _has_merge_E_to_G_one_row(ws, stats_row):
        ws.merge_cells(start_row=stats_row, start_column=5, end_row=stats_row, end_column=7)
    # Chỉ gán ô góc trái (E); F/G trong merge là MergedCell, không gán được.
    ws.cell(stats_row, 5, value="BÁO CÁO THỐNG KÊ")


def _stats_counts(rows):
    order = ["Tối khẩn", "Khẩn", "Cao", "Trung bình", "Thấp"]
    counts_pri = {k: 0 for k in order}
    for r in rows:
        lab = (r.get("phan_loai") or "").strip()
        if lab in counts_pri:
            counts_pri[lab] += 1
    return order, counts_pri


def write_bao_cao_from_template(
    path: str,
    template_path: str,
    rows: list,
    week_monday: datetime,
    week_sunday: datetime,
):
    """Sao chép file mẫu (giữ ảnh/logo + màu theme), điền dữ liệu và thống kê."""
    shutil.copyfile(template_path, path)
    wb = load_workbook(path)
    ws = wb.active

    orig_stats_row = _find_stats_title_row(ws)
    if not orig_stats_row:
        wb.close()
        raise ValueError("Không tìm thấy khối BÁO CÁO THỐNG KÊ trong file mẫu.")

    # Giữ nguyên toàn bộ khối thống kê (bao gồm “THỐNG KÊ THEO NỘI DUNG”).
    # Chỉ thay thế vùng dữ liệu từ dòng 5 đến trước “BÁO CÁO THỐNG KÊ”.
    template_data_rows = orig_stats_row - 5
    if template_data_rows < 0:
        template_data_rows = 0

    # Lưu style từ mẫu ngay trong workbook này (trước khi xóa/chèn dòng)
    style_src = [ws.cell(5, c) for c in range(1, 11)]
    style_pri_e = ws.cell(orig_stats_row + 2, 5)
    style_pri_g = ws.cell(orig_stats_row + 2, 7)
    ref_title = ws.cell(orig_stats_row, 5)
    title_font = copy(ref_title.font)
    title_fill = copy(ref_title.fill)
    title_align = copy(ref_title.alignment)
    title_border = copy(ref_title.border)

    if template_data_rows:
        ws.delete_rows(5, template_data_rows)
    
    # Thêm 3 dòng trống để cách ra so với bảng trên
    spacing_rows = 3
    insert_count = len(rows) + spacing_rows
    if insert_count > 0:
        ws.insert_rows(5, insert_count)

    new_stats_row = orig_stats_row - template_data_rows + insert_count

    _ensure_bao_cao_thong_ke_merge(ws, new_stats_row)
    c_title = ws.cell(new_stats_row, 5)
    c_title.font = title_font
    c_title.fill = title_fill
    c_title.alignment = title_align
    c_title.border = title_border
    _cleanup_orphan_EG_row_merges(ws, new_stats_row)

    header_text = _week_header_text(week_monday, week_sunday)
    ws["A3"].value = header_text

    for i, rec in enumerate(rows, start=1):
        rr = 4 + i  # dữ liệu bắt đầu từ dòng 5
        values = [
            i,
            rec.get("dich_vu") or "",
            rec.get("ngay") or "",
            rec.get("phan_loai") or "",
            rec.get("tieu_de") or "",
            rec.get("chi_tiet") or "",
            rec.get("muc_do") or "",
            rec.get("phuong_an") or "",
            rec.get("ket_qua") or "",
            rec.get("de_xuat") or "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(rr, col, value=val)
            _apply_cell_style(style_src[col - 1], cell)

    hdr = new_stats_row + 1
    order, counts_pri = _stats_counts(rows)

    for j, lab in enumerate(order):
        r = hdr + 1 + j
        ce = ws.cell(r, 5, value=lab)
        cg = ws.cell(r, 7, value=counts_pri[lab])
        _apply_cell_style(style_pri_e, ce)
        _apply_cell_style(style_pri_g, cg)

    # Đảm bảo kẻ viền (border) cho toàn bộ 9 dòng của bảng thống kê (cột E, F, G)
    # Bao gồm 5 dòng phân loại lỗi và 4 dòng "Sự vụ VAS, VNPost..." ở dưới (vùng tô vàng)
    for j in range(1, 10):
        r = hdr + j
        for col in [5, 6, 7]:
            ws.cell(r, col).border = copy(style_pri_e.border)

    # Phần “THỐNG KÊ THEO NỘI DUNG” giữ nguyên text theo file mẫu để sửa tay.

    _embed_logo_from_template(ws, template_path)
    wb.save(path)
    wb.close()


def write_bao_cao_workbook_plain(path, rows, week_monday: datetime, week_sunday: datetime):
    """Không có file mẫu: tạo Excel mới (màu cơ bản như trước)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BC_T1"

    for col, w in enumerate(
        [13.0, 19.44, 22.33, 13.0, 23.0, 31.89, 34.66, 34.89, 32.11, 13.0], start=1
    ):
        ws.column_dimensions[get_column_letter(col)].width = w

    base = Font(name=FONT_NAME, size=FONT_SIZE, bold=False)
    bold = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_top_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

    header_block = _week_header_text(week_monday, week_sunday)
    meta_doc = (
        "Mã số: BM-VINATTI-17-01\n"
        "Ngày ban hành: 13/05/2025\n"
        "Lần ban hành:01"
    )

    ws.merge_cells("C2:D2")
    ws.merge_cells("E2:G2")
    ws["E2"] = "BÁO CÁO TUẦN"
    ws["E2"].font = bold
    ws["E2"].alignment = center_wrap
    ws["H2"] = meta_doc
    ws["H2"].font = bold
    ws["H2"].alignment = left_wrap

    ws.merge_cells("A3:I3")
    ws["A3"] = header_block
    ws["A3"].font = bold
    ws["A3"].alignment = center_wrap

    headers = [
        "STT",
        "DỊCH VỤ",
        "NGÀY",
        "PHÂN LOẠI LỖI",
        "TIÊU ĐỀ",
        "CHI TIẾT LỖI",
        "MỨC ĐỘ ẢNH HƯỞNG",
        "PHƯƠNG ÁN XỬ LÝ",
        "KẾT QUẢ XỬ LÝ",
        "ĐỀ XUẤT",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = bold
        cell.alignment = center_wrap

    data_start = 5
    for i, r in enumerate(rows, start=1):
        row_idx = data_start + i - 1
        values = [
            i,
            r.get("dich_vu") or "",
            r.get("ngay") or "",
            r.get("phan_loai") or "",
            r.get("tieu_de") or "",
            r.get("chi_tiet") or "",
            r.get("muc_do") or "",
            r.get("phuong_an") or "",
            r.get("ket_qua") or "",
            r.get("de_xuat") or "",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = base
            c.alignment = center_top_wrap if col <= 4 else left_wrap

    last_data_row = data_start + max(len(rows), 0) - 1
    if not rows:
        last_data_row = data_start - 1

    stats_title_row = last_data_row + 3
    merge_stats = f"E{stats_title_row}:G{stats_title_row}"
    ws.merge_cells(merge_stats)
    ws.cell(row=stats_title_row, column=5, value="BÁO CÁO THỐNG KÊ").font = bold
    ws.cell(row=stats_title_row, column=5).alignment = center_wrap

    hdr = stats_title_row + 1
    ws.cell(row=hdr, column=5, value="THỐNG KÊ THEO PHÂN LOẠI LỖI").font = bold
    ws.cell(row=hdr, column=5).alignment = center_wrap
    ws.cell(row=hdr, column=6, value="THỐNG KÊ THEO NỘI DUNG").font = bold
    ws.cell(row=hdr, column=6).alignment = center_wrap
    ws.cell(row=hdr, column=7, value="SỐ LƯỢNG").font = bold
    ws.cell(row=hdr, column=7).alignment = center_wrap

    order, counts_pri = _stats_counts(rows)

    row_ptr = hdr + 1
    for lab in order:
        ws.cell(row=row_ptr, column=5, value=lab).font = base
        ws.cell(row=row_ptr, column=5).alignment = center_wrap
        ws.cell(row=row_ptr, column=7, value=counts_pri[lab]).font = base
        ws.cell(row=row_ptr, column=7).alignment = center_wrap
        row_ptr += 1

    # Phần “THỐNG KÊ THEO NỘI DUNG” để trống (tự điền tay nếu cần).

    wb.save(path)


def write_bao_cao_workbook(path, rows, week_monday: datetime, week_sunday: datetime):
    template_path = os.environ.get("BC_TUAN_TEMPLATE", DEFAULT_TEMPLATE)
    if os.path.isfile(template_path):
        try:
            write_bao_cao_from_template(path, template_path, rows, week_monday, week_sunday)
            return
        except Exception as exc:
            print(f"[BC-TUAN] Không dùng được file mẫu ({template_path}): {exc}")
    write_bao_cao_workbook_plain(path, rows, week_monday, week_sunday)


def tao_bao_cao_tuan():
    now = datetime.now()
    week_start, week_end = week_bounds(now)

    api_token = os.environ.get(
        "QUIK_API_TOKEN",
        "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VySWQiOjU2LCJlbWFpbCI6ImtoYW5odGRAdmluYXR0aS52biIsInJvbGUiOiJlbXBsb3llZSIsImlhdCI6MTc3NjQ3ODA3MiwiZXhwIjoxNzc3MDgyODcyfQ.wTc8eBkHb6GHIZywzEBa_pH-k95jFAAf8c9eiYNRyGw",
    )
    url = "https://quik.works/api/incidents"
    headers = {
        "Authorization": f"Bearer {api_token}",
        "Accept": "application/json",
    }

    response = requests.get(url, headers=headers, timeout=120)
    response.raise_for_status()
    data = response.json()
    if isinstance(data, dict) and "data" in data:
        data = data["data"]

    result = []
    for item in data:
        created_at = item.get("createdAt", "")
        if not in_week(created_at, week_start, week_end):
            continue

        completed_at = item.get("completedAt", "")

        incident_id = item.get("id")
        detail_url = f"https://quik.works/api/incidents/{incident_id}"
        detail_resp = requests.get(detail_url, headers=headers, timeout=120)
        detail_resp.raise_for_status()
        detail_data = detail_resp.json()

        if isinstance(detail_data, dict) and "data" in detail_data:
            detail_data = detail_data["data"]

        affected_services = detail_data.get("affectedServices", [])
        final_service_name = ""
        if isinstance(affected_services, list):
            names = [
                srv.get("name", "")
                for srv in affected_services
                if isinstance(srv, dict)
            ]
            final_service_name = ", ".join(names)
        elif isinstance(affected_services, dict):
            final_service_name = affected_services.get("name", "")

        translated_priority = priority_label(item.get("priority"))

        muc_do = first_nonempty_str(
            detail_data,
            item,
            keys=(
                "businessImpact",
                "impactDescription",
                "impact",
                "customerImpact",
                "severityNote",
                "affectedScope",
            ),
        )
        phuong_an = first_nonempty_str(
            detail_data,
            item,
            keys=(
                "resolutionPlan",
                "mitigation",
                "workaround",
                "actionPlan",
                "handlingMeasures",
                "correctiveAction",
                "remediation",
            ),
        )

        result.append(
            {
                "dich_vu": final_service_name,
                "ngay": format_ngay_cell(created_at, completed_at),
                "phan_loai": translated_priority,
                "tieu_de": item.get("title") or "",
                "chi_tiet": item.get("description") or "",
                "muc_do": muc_do,
                "phuong_an": phuong_an,
                "ket_qua": item.get("completionNote") or "",
                "de_xuat": "",
            }
        )

    filename = f"BC-TUAN{now.strftime('%d.%m.%y')}.xlsx"
    write_bao_cao_workbook(filename, result, week_start, week_end)
    return filename


@bot.message_handler(commands=["start", "help"])
def handle_start(message):
    bot.reply_to(
        message,
        "Chào bạn. Gửi lệnh /baocao để tạo và nhận file báo cáo tuần (Excel).",
    )


@bot.message_handler(commands=["baocao"])
def handle_baocao(message):
    bot.reply_to(message, "Em đang tổng hợp báo cáo, anh Chuẩn chờ em chút ạaaa...")
    filename = None
    try:
        filename = tao_bao_cao_tuan()
        with open(filename, "rb") as file:
            bot.send_document(message.chat.id, file)
    except Exception as e:
        bot.reply_to(message, f"Có lỗi xảy ra: {e}")
    finally:
        if filename:
            try:
                os.remove(filename)
            except OSError:
                pass


if __name__ == "__main__":
    print("Bot Telegram đang chạy. Nhấn Ctrl+C để dừng.")
    bot.infinity_polling(skip_pending=True, timeout=60, long_polling_timeout=60)
